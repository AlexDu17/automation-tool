"""读取输入 Excel 文件，兼容两种情况：

1. 真正的 .xlsx（Office Open XML，zip 容器）—— 用 openpyxl 读取。
2. 文件名是 .xlsx 但内容其实是老式 .xls（BIFF/OLE 二进制格式，很多学校
   教务系统导出的成绩表都是这种情况）—— 用 xlrd 读取。

统一返回一个二维列表 List[List[Any]]，所有行补齐到相同列数，方便后续
按下标解析表头，不用关心具体是哪种引擎读出来的。
"""
from __future__ import annotations

from typing import Any, List

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
import xlrd
import zipfile


class UnsupportedWorkbookError(ValueError):
    """输入文件不是可识别的 Excel 格式。"""


def _pick_sheet_name(names: List[str]) -> str:
    for name in names:
        if "成绩" in name:
            return name
    return names[0]


def _rows_from_openpyxl(path: str) -> List[List[Any]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheet_name = _pick_sheet_name(wb.sheetnames)
        ws = wb[sheet_name]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    return rows


def _rows_from_xlrd(path: str) -> List[List[Any]]:
    book = xlrd.open_workbook(path)
    sheet_name = _pick_sheet_name(book.sheet_names())
    sheet = book.sheet_by_name(sheet_name)
    rows = [sheet.row_values(r) for r in range(sheet.nrows)]
    return rows


def _normalize(rows: List[List[Any]]) -> List[List[Any]]:
    width = max((len(r) for r in rows), default=0)
    normalized = []
    for row in rows:
        cells = list(row) + [None] * (width - len(row))
        cleaned = []
        for cell in cells:
            if isinstance(cell, str):
                cell = cell.strip()
                if cell == "":
                    cell = None
            cleaned.append(cell)
        normalized.append(cleaned)
    return normalized


def load_raw_rows(path: str) -> List[List[Any]]:
    """读取 path 指向的 Excel 文件，返回补齐后的二维表格。"""
    try:
        rows = _rows_from_openpyxl(path)
    except (InvalidFileException, zipfile.BadZipFile, KeyError):
        try:
            rows = _rows_from_xlrd(path)
        except Exception as exc:  # noqa: BLE001 - 转成统一的错误类型
            raise UnsupportedWorkbookError(
                f"无法识别的 Excel 文件格式：{path}"
            ) from exc
    return _normalize(rows)
