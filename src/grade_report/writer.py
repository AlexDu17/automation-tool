"""把统计结果写成一个多 Sheet 的 Excel 工作簿。"""
from __future__ import annotations

from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import analysis as A
from .model import ParsedSheet, Student

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
OVERALL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HIGHEST_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
LOWEST_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")

METRIC_LABELS = ["A+", "A", "A以上", "B+", "B+以上", "B", "B以上", "C+", "C"]


def _style_header_row(ws: Worksheet, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER


def _autofit(ws: Worksheet, ncols: int, min_width: int = 8, max_width: int = 22) -> None:
    widths = [min_width] * ncols
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or cell.column > ncols:
                continue
            length = len(str(cell.value))
            idx = cell.column - 1
            widths[idx] = max(widths[idx], min(length + 2, max_width))
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_grade_distribution(ws: Worksheet, students: List[Student], subjects: List[str]) -> None:
    row = 1
    max_col = 1 + len(METRIC_LABELS) * 2
    for subject in subjects:
        ws.cell(row=row, column=1, value=f"{subject}等级人数及占比")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        ws.cell(row=row, column=1).font = HEADER_FONT
        row += 1

        ws.cell(row=row, column=1, value=subject)
        col = 2
        for label in METRIC_LABELS:
            ws.cell(row=row, column=col, value=label)
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
            col += 2
        _style_header_row(ws, row, max_col)
        header_label_row = row
        row += 1

        col = 2
        for _ in METRIC_LABELS:
            ws.cell(row=row, column=col, value="人数")
            ws.cell(row=row, column=col + 1, value="占比")
            col += 2
        _style_header_row(ws, row, max_col)
        row += 1

        dist_rows = A.build_grade_distribution(students, subject)
        for dist_row in dist_rows:
            ws.cell(row=row, column=1, value=dist_row.label)
            col = 2
            for count, ratio in A.distribution_metric_values(dist_row.counts, dist_row.total):
                ws.cell(row=row, column=col, value=count)
                ratio_cell = ws.cell(row=row, column=col + 1, value=ratio)
                if ratio is not None:
                    ratio_cell.number_format = "0.0%"
                col += 2
            if dist_row.label == A.OVERALL_LABEL:
                for c in range(1, max_col + 1):
                    ws.cell(row=row, column=c).fill = OVERALL_FILL
            row += 1

        row += 1  # 空行分隔

    ws.freeze_panes = "A4"
    _autofit(ws, max_col)


def write_subject_averages(ws: Worksheet, students: List[Student], subjects: List[str]) -> None:
    max_col = 1 + len(subjects) * 3
    col = 2
    for subject in subjects:
        ws.cell(row=1, column=col, value=subject)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        col += 3
    _style_header_row(ws, 1, max_col)

    col = 2
    for _ in subjects:
        ws.cell(row=2, column=col, value="教师")
        ws.cell(row=2, column=col + 1, value="均分")
        ws.cell(row=2, column=col + 2, value="排名")
        col += 3
    _style_header_row(ws, 2, max_col)

    per_subject_rows = {subject: A.build_subject_averages(students, subject) for subject in subjects}
    n_rows = len(next(iter(per_subject_rows.values())))
    for i in range(n_rows):
        r = 3 + i
        first_row = next(iter(per_subject_rows.values()))[i]
        ws.cell(row=r, column=1, value=first_row.label)
        col = 2
        for subject in subjects:
            avg_row = per_subject_rows[subject][i]
            ws.cell(row=r, column=col, value=None)  # 教师：留空，老师手动填写
            avg_cell = ws.cell(row=r, column=col + 1, value=round(avg_row.average, 2) if avg_row.average is not None else None)
            avg_cell.number_format = "0.00"
            ws.cell(row=r, column=col + 2, value=avg_row.rank)
            col += 3
        if first_row.label == A.OVERALL_LABEL:
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = OVERALL_FILL

    ws.freeze_panes = "A3"
    _autofit(ws, max_col)


def write_rank_segment(ws: Worksheet, students: List[Student], subject: str) -> None:
    class_codes, rows = A.build_rank_segments(students, subject)
    max_col = 1 + len(class_codes) * 2

    ws.cell(row=1, column=1, value="分段")
    col = 2
    for class_code in class_codes:
        ws.cell(row=1, column=col, value=class_code)
        ws.cell(row=1, column=col + 1, value=subject)
        col += 2
    _style_header_row(ws, 1, max_col)

    for i, seg_row in enumerate(rows):
        r = 2 + i
        ws.cell(row=r, column=1, value=seg_row.label)
        col = 2
        avg_cols_by_class = {}
        for class_code in class_codes:
            first_val, second_val = seg_row.cells[class_code]
            ws.cell(row=r, column=col, value=first_val)
            ws.cell(row=r, column=col + 1, value=second_val)
            avg_cols_by_class[class_code] = col
            col += 2
        if seg_row.is_average_row:
            averages = {
                class_code: seg_row.cells[class_code][0]
                for class_code in class_codes
                if seg_row.cells[class_code][0] is not None
            }
            if averages:
                highest = max(averages.values())
                lowest = min(averages.values())
                for class_code, value in averages.items():
                    fill = None
                    if value == highest:
                        fill = HIGHEST_FILL
                    elif value == lowest:
                        fill = LOWEST_FILL
                    if fill is not None:
                        ws.cell(row=r, column=avg_cols_by_class[class_code]).fill = fill

    ws.freeze_panes = "A2"
    _autofit(ws, max_col)


def write_raw_scores(ws: Worksheet, students: List[Student], parsed: ParsedSheet) -> None:
    base_names = list(parsed.base_columns.keys())
    subject_names = [s.display_name for s in parsed.subjects]
    headers = list(base_names)
    for name in subject_names:
        headers.append(f"{name}分数")
        headers.append(f"{name}等级")
    headers.append("年级排名")
    max_col = len(headers)

    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    _style_header_row(ws, 1, max_col)

    ranked = A.build_raw_scores(students, parsed)
    for i, (student, rank) in enumerate(ranked):
        r = 2 + i
        col = 1
        for name in base_names:
            ws.cell(row=r, column=col, value=student.row.get(name))
            col += 1
        for name in subject_names:
            score, grade = student.scores.get(name, (None, None))
            ws.cell(row=r, column=col, value=score)
            col += 1
            ws.cell(row=r, column=col, value=grade)
            col += 1
        ws.cell(row=r, column=col, value=rank)

    ws.freeze_panes = "A2"
    _autofit(ws, max_col)


def write_class_roster(ws: Worksheet, students: List[Student], class_code: str, parsed: ParsedSheet) -> None:
    subject_names = [s.display_name for s in parsed.subjects]
    headers = ["姓名", "行政班级"] + [f"{name}{suffix}" for name in subject_names for suffix in ("分数", "等级")]
    max_col = len(headers)

    for col, name in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=name)
    _style_header_row(ws, 1, max_col)

    roster = A.build_class_roster(students, class_code, parsed)
    for i, student in enumerate(roster):
        r = 2 + i
        ws.cell(row=r, column=1, value=student.row.get("姓名"))
        ws.cell(row=r, column=2, value=student.class_code)
        col = 3
        for name in subject_names:
            score, grade = student.scores.get(name, (None, None))
            ws.cell(row=r, column=col, value=score)
            col += 1
            ws.cell(row=r, column=col, value=grade)
            col += 1

    ws.freeze_panes = "A2"
    _autofit(ws, max_col)


def build_workbook(students: List[Student], parsed: ParsedSheet) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    dist_subjects = A.distribution_subjects(parsed)
    rank_subjects = A.rank_sheet_subjects(parsed)
    class_codes = A.sorted_class_codes(students)

    ws = wb.create_sheet("等级分布")
    write_grade_distribution(ws, students, dist_subjects)

    ws = wb.create_sheet("各科平均分")
    write_subject_averages(ws, students, [s for s in dist_subjects if s != parsed.total_display_name])

    main_rank_subjects = [s for s in rank_subjects if s not in parsed.component_display_names]
    component_subjects = [s for s in rank_subjects if s in parsed.component_display_names]

    for subject in main_rank_subjects:
        ws = wb.create_sheet(f"{subject}成绩")
        write_rank_segment(ws, students, subject)

    ws = wb.create_sheet("原始成绩")
    write_raw_scores(ws, students, parsed)

    for subject in component_subjects:
        ws = wb.create_sheet(f"{subject}成绩")
        write_rank_segment(ws, students, subject)

    for class_code in class_codes:
        ws = wb.create_sheet(class_code)
        write_class_roster(ws, students, class_code, parsed)

    return wb
